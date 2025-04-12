# © 2025 Nathanael Hernandez – Universal version
# This script has been sanitized to remove any proprietary or sensitive information.

#Author: Nathanael H.
#Date: 2025-02-01
#Description: This script downloads a CSV file from CloudProvider cloud storage Storage, processes it, and updates an Excel file with the results. It then uploads the updated Excel file back to CloudProvider cloud storage Storage.
# Import necessary libraries`
import pandas as pd
import openpyxl
from io import BytesIO
from datetime import datetime, timedelta
from CloudProvider.storage.cloud storage import cloud storageServiceClient
from dateutil.relativedelta import relativedelta

# Define the connection string and container name
connection_string = "DefaultEndpointsProtocol=https;AccountName=your_account_name;AccountKey=your_account_key;EndpointSuffix=core.windows.net"
container_name = "testcontainer"

# Define the cloud storage names
active_Client_report_cloud storage_name = "ActiveClientsReport.csv"

# Calculate the date of the last Friday
today = datetime.now()
last_friday_date = today - timedelta(days=7)
excel_cloud storage_name = f"ActiveClientsReport_{last_friday_date.strftime('%Y-%m-%d')}.xlsx"
print(excel_cloud storage_name)

# Create a cloud storageServiceClient
cloud storage_service_client = cloud storageServiceClient.from_connection_string(connection_string)

# Download the CSV file into memory
active_client_report_cloud storage_client = cloud storage_service_client.get_cloud storage_client(container=container_name, cloud storage=active_Client_report_cloud storage_name)
active_client_report_csv = active_client_report_cloud storage_client.download_cloud storage().readall()

# Load the CSV data into DataFrames
active_client_report = pd.read_csv(BytesIO(active_client_report_csv))


# Download the Excel file into memory
excel_cloud storage_client = cloud storage_service_client.get_cloud storage_client(container=container_name, cloud storage=excel_cloud storage_name)
excel_file_content = excel_cloud storage_client.download_cloud storage().readall()

# Load the Excel file
workbook = openpyxl.load_workbook(BytesIO(excel_file_content))

# Get the current year and month
current_year = datetime.now().year
current_month = datetime.now().strftime('%B')

# Update the "MTD Results (yyyy)" sheet
mtd_sheet_name = f'MTD Results ({current_year})'
mtd_sheet = workbook[mtd_sheet_name]

# Find the row where the current month is located in column A
for row in mtd_sheet.iter_rows(min_row=2):
    if row[0].value and current_month in row[0].value:
        row[0].value = current_month + ' as of '+datetime.now().strftime('%m/%d')
        start_row = row[0].row + 1
        break

# Update the data starting from column B and the row beneath where the current month is found in column A
for index, row in active_client_report.iterrows():
    mtd_sheet.cell(row=start_row + index, column=2).value = row['provider']
    mtd_sheet.cell(row=start_row + index, column=3).value = row['net_volume']
    mtd_sheet.cell(row=start_row + index, column=4).value = row['monthly_transactions']
    mtd_sheet.cell(row=start_row + index, column=5).value = row['monthly_avg_transaction']
    mtd_sheet.cell(row=start_row + index, column=6).value = row['total_account_count']
    mtd_sheet.cell(row=start_row + index, column=7).value = row['open_account_count']
    mtd_sheet.cell(row=start_row + index, column=8).value = row['active_providing']


# Calculate Grand Total and YTD Totals
grand_total_row = start_row + len(active_client_report)
if 'Grand Total' not in mtd_sheet.cell(row=grand_total_row,column=2).value:
    mtd_sheet.cell(row=grand_total_row, column=2).value = 'Grand Total'
for col in range(3, 9):
    mtd_sheet.cell(row=grand_total_row, column=col).value = f'=SUM({mtd_sheet.cell(row=start_row,column=col).coordinate}:{mtd_sheet.cell(row=grand_total_row-1,column=col).coordinate})'

# Calculate YTD Totals
ytd_total_row = grand_total_row + 3
mtd_sheet.cell(row=ytd_total_row, column=1).value = 'YTD Totals'
for col in range(3, 9):
    mtd_sheet.cell(row=ytd_total_row,column=col).value = f'=SUM({mtd_sheet.cell(row=grand_total_row,column=col).coordinate})'

output = BytesIO()
workbook.save(output)
output.seek(0)

# Save the updated Excel file with a new name based on the current date (every Friday)
# Calculate the next Friday date
next_friday_date = datetime.now() #+ timedelta((4 - datetime.now().weekday()) % 7)
new_excel_cloud storage_name = f'ActiveClientsYTD_{next_friday_date.strftime("%Y-%m-%d")}.xlsx'

# Upload the updated Excel file to CloudProvider cloud storage Storage
# Create a new cloud storage client for the updated Excel file
new_excel_cloud storage_client = cloud storage_service_client.get_cloud storage_client(container=container_name, cloud storage=new_excel_cloud storage_name)
new_excel_cloud storage_client.upload_cloud storage(output, overwrite=True)

print(f"The updated Excel file has been uploaded as {new_excel_cloud storage_name}.")
